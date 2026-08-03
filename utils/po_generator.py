from pathlib import Path
from openpyxl import load_workbook
import io
import pandas as pd
from datetime import datetime
from copy import copy
from openpyxl.utils.cell import range_boundaries

SCRIPT_DIR = Path(__file__).parent
TEMPLATE_PATH = SCRIPT_DIR / 'po_templates' / 'po_template.xlsx'

MASTER_MAPPING_PO = {
    "cvn_number_cell": "B4",
    "model_name_cell": "J5", # Model #
    "model_code_cell": "J6",
    "total_vehicles_cell": "P6",    # Total cars on the job
    "accessory_start_row": 10,
    "accessory_code_col" : 2, # Column B
    "accessory_qty_col": 3, # Column C
    "accessory_desc_col" :  5, # Column E
    "accessory_fitting_time_col" :  12, # Column L
}
# Supplier's info:
OPERADORA_PAJARITO = {'name':'Operadora Pajarito',
                      'address':'Keizershoek 170, \n 2550 Kontich',
                      'price': 67}

TALLER_DON_JOSE = {'name':'Taller Don Jose',
                      'address':'Rupelmondestraat 17, \n 9150 Bazel',
                      'price': 70}

# Function to select supplier data from database
def vendor_data(subcontractor_name: str) -> dict:
    ''' Reads subcontractors data and extract the data for the selected supplier
    Args
        subcontractor_name - str: name of the selected subcontractor

    Returns
        vendor_data - dict: all de details for the selected supplier

    '''
    subcontractors = pd.read_excel(SCRIPT_DIR.parent / 'data' / 'subcontractors.xlsx', sheet_name='Ext_services')
    subcontractor_data = subcontractors[subcontractors['name'] == subcontractor_name]

    print(subcontractor_data)
    return subcontractor_data


# Function to copy/paste the formating of a row
def copy_row_format(ws, source_row: int, target_row: int):
    """Copies all cell formatting from a source row to a target row."""
    # Loop over every active column in the spreadsheet
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        # Explicitly copy styles over to the new cell if they exist
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.alignment = copy(source_cell.alignment)



# Function to add new rows  
def adjust_accessory_rows(ws, num_rows: int, default_rows: int = 4, base_row: int = 23):
    """Inserts extra rows, fixes broken merges, and clones exact merge formats."""
    if num_rows <= default_rows:
        return

    rows_to_add = num_rows - default_rows
    template_height = ws.row_dimensions[base_row].height

    # 1. Identify horizontal merge patterns on our template row (Row 23)
    # Format saved as a list of tuples: (start_column, end_column)
    row_23_merges = []
    
    # We copy the current merges list because we will modify ws.merged_cells during adjustment
    current_merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in current_merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        
        # If the merge is exactly on row 23, remember its column boundaries
        if min_row == base_row and max_row == base_row:
            row_23_merges.append((min_col, max_col))
        
        # FIX FOR DISTORTION: If the merge is below row 23, openpyxl stretches it incorrectly.
        # We manually unmerge it before row insertion, then fix it later.
        elif min_row > base_row:
            ws.unmerge_cells(str(merged_range))
            merged_range.shift(row_shift=rows_to_add)
            ws.merged_cells.add(merged_range)

    # 2. Safely insert the blank rows now that bottom merges are protected
    ws.insert_rows(base_row + 1, amount=rows_to_add)

    # 3. Apply styles, heights, and identical horizontal merges to new rows
    for i in range(1, rows_to_add + 1):
        target_row = base_row + i
        
        # Apply standard cell styling and height
        copy_row_format(ws, source_row=base_row, target_row=target_row)
        ws.row_dimensions[target_row].height = template_height
        
        # Apply the exact horizontal merge mapping cloned from row 23
        for start_col, end_col in row_23_merges:
            ws.merge_cells(
                start_row=target_row, 
                start_column=start_col, 
                end_row=target_row, 
                end_column=end_col
            )


# Function that reads the worsheet and returns the useful information
def collect_data(uploaded_ws, mapping: dict) -> pd.DataFrame:
    """
    Reads data from the worksheet, extracts relevant information, 
    and returns a DataFrame containing the base df for the scan list.
    
    Args:
        uploaded_ws (Path): Path to the uploaded worksheet file.

    Returns:
        base_scan_list: DataFrame containing the base scan list with columns:
    """
    # 1. Open the uploaded master worksheet
    # data_only=True ensures we read final values, not active Excel formulas
    workbook = load_workbook(uploaded_ws, data_only=True)
    worksheet = workbook["Worksheet"] #Target the sheet that contains the data
   
    # 2. Extract the header details from the top cells
    collected_data = {
        'cvn_num' : worksheet[mapping['cvn_number_cell']].value or 'NO_REF_FOUND',
        'vehicles_qty' : worksheet[mapping["total_vehicles_cell"]].value or 0,
        'model_name' : worksheet[mapping['model_name_cell']].value or 'NO_MODEL_FOUND',
        'model_code' : worksheet[mapping['model_code_cell']].value or 'NO_MODEL_CODE_FOUND'
        # Add mapping to CPS Stickers
    }

    # 3. Read worksheet, find and count number of accessories and description
    current_row = MASTER_MAPPING_PO["accessory_start_row"] #Start checking on row 10
    accessories_dict = {}  # Dictionary to hold accessory codes and their descriptions
    blank_rows = 0

    while True:
        # 1. Read the value from B2 and E10
        code_value = worksheet.cell(row=current_row, column=mapping["accessory_code_col"]).value #mapping cell B10
        accessory_desc = worksheet.cell(row=current_row, column=mapping["accessory_desc_col"]).value #mapping cell E10
        qty_vin = worksheet.cell(row=current_row, column=mapping["accessory_qty_col"]).value #mapping cell C10
        fitting_time = worksheet.cell(row=current_row, column=mapping["accessory_fitting_time_col"]).value or 0 #mapping cell L10

        # 2. Stop if two consecutive rows are blank
        if blank_rows >= 2:  
                break
        
        # 3. Handle blank cells  
        if code_value is None or str(code_value).strip() == '':
            blank_rows += 1
            current_row += 1
            continue  

        # 5. Process valid rows
        else:
            blank_rows = 0  # Reset blank row counter
            accessories_dict[code_value] = {
                "description": accessory_desc,
                "quantity": qty_vin,
                "fitting_time": fitting_time
            }
            current_row += 1
   
        num_rows = len(accessories_dict)

    return collected_data, accessories_dict, num_rows


def fill_template(TEMPLATE_PATH: str, supplier: str, collected_data: dict, accessories_dict: dict, num_rows: int):
    '''
    Fills the template based on the supplier and returns the filled document.
    
    Args:
        supplier (str): The selected supplier.
        collected_data (dict): Dictionary containing header details.
        accessories_dict (dict): Dictionary containing accessory details.
        num_rows (int): Number of accessory rows.
    Returns:
        filled_template: The filled template document.
    '''
    # 1. Load the template
    po_temp_wb = load_workbook(TEMPLATE_PATH, data_only=True)
    po_template = po_temp_wb.active

    adjust_accessory_rows(po_template, num_rows, default_rows=4, base_row=23)

    # 2. Select supplier  
    if supplier == 'Operadora Pajarito':
        supplier = OPERADORA_PAJARITO
    elif supplier == 'Taller Don Jose':
        supplier = TALLER_DON_JOSE
    
    # 2. Fill in the header details
    po_template['B6'] = supplier['name']
    po_template['B7'] = supplier['address']
    po_template['L4'] = datetime.now().strftime('%d/%m/%Y')
    po_template['L8'] = collected_data['cvn_num']
    po_template['B13'] = f'{collected_data['vehicles_qty']}x {collected_data['model_name']} \n {collected_data['model_code']}'
    po_template['K19'] = supplier['price']

    # Now fill accessories list as of row 20
    start_row = 20

    # item_data = dict of each accessory. Eg 3585 : {'description': 'Pintle hook'}
    for idx, (key, item_data) in enumerate(accessories_dict.items()):
        current_target_row = start_row + idx

        # Define variable to be able to refer to them later.
        # At least supplier fitting_time needs a variable name to then be added
        acc_code = key # Column B
        acc_desc = item_data.get('description', 'Not found')    # Column C
        qty_vin =  item_data.get('quantity', 0)
        fitting_time = item_data.get('fitting_time')  # Column J
        supplier_price = supplier['price'] 
        price_vin = fitting_time * supplier_price
        total = fitting_time * supplier_price * collected_data['vehicles_qty']
        
        # Write to specific columns based on your template's layout
        # (Remember: use the top-left cell coordinate if the column is merged!)
        po_template.cell(row=current_target_row, column=2).value = acc_code
        po_template.cell(row=current_target_row, column=3).value = acc_desc
        po_template.cell(row=current_target_row, column=9).value = qty_vin     # Column I
        po_template.cell(row=current_target_row, column=10).value = fitting_time
        po_template.cell(row=current_target_row, column=11).value = price_vin
        po_template.cell(row=current_target_row, column=12).value = total 

        # Fix unit price if fitting time == '-'
        if type(price_vin) == str:
            price_vin = '-'
        else:
            price_vin = price_vin
        # Fix total price if fitting time == '-'
        if type(total) == str:
            total  = '-'
        else:
            total = total

    # Convert to 
    po_stream = io.BytesIO()
    po_temp_wb.save(po_stream)
    po_stream.seek(0)

    return po_stream


