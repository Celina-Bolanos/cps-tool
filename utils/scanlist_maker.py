from openpyxl import load_workbook
import pandas as pd
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
ACCESSORIES_DB = SCRIPT_DIR / 'databases' / 'ACCESSORIESCODES_DATABASE.csv'
ITEMS_DB = SCRIPT_DIR / 'databases' / 'ITEMS_DATABASE.csv'

MASTER_MAPPING = {
    "cvn_number_cell": "B4",
    "total_vehicles_cell": "P6",    # Total cars on the job
    "accessory_start_row": 10,
    "accessory_code_col" : 2, # Column B
    'total_qty_col' : 4 # Col D
}

# Function that reads the worsheet and returns the useful information
def read_ws(uploaded_ws, mapping: dict) -> pd.DataFrame:
    """
    Reads data from the worksheet, extracts relevant information, 
    and returns a DataFrame containing the base df for the scan list.
    
    Args:
        uploaded_ws (Path): Path to the uploaded worksheet file.

    Returns:
        base_scan_list: DataFrame containing the base scan list with columns:
    """
    # 1. Open the uploaded master worksheet
    # data_only=True ensures we read final values, not active Excel formulas
    workbook = load_workbook(uploaded_ws, data_only=True)
    worksheet = workbook["Worksheet"] #Target the sheet that contains the data
   
    # 2. Extract the header details from the top cells
    collected_data = {
        'cvn_num' : worksheet[mapping['cvn_number_cell']].value or 'NO_REF_FOUND',
        'vehicles_qty' : worksheet[mapping["total_vehicles_cell"]].value or 0,
    }

    # Read worksheet, find and count number of accessories
    current_row = MASTER_MAPPING["accessory_start_row"] #Start checking on row 10
    accessory_codes_dict = {}
    blank_rows = 0

    while True:
        # 1. Read the value from Column B2
        code_value = worksheet.cell(row=current_row, column=mapping["accessory_code_col"]).value #mapping cell B10
        total_qty = worksheet.cell(row=current_row, column=mapping['total_qty_col']).value # Poit to cell D10

        # 2. Stop if two consecutive rows are blank
        if blank_rows >= 2:  
                break

        # 3. Skip Ext. service lines
        if str(code_value).strip() in ['Ext. Service', 'ext. Service']:
            current_row += 1
            continue
        
        # 4. Handle blank cells  
        elif code_value is None or str(code_value).strip() == '':
            blank_rows += 1
            current_row += 1  # FIX: Move to next row so you don't get stuck
            continue  

        # 5. Process valid rows
        else:
            blank_rows = 0  # Reset blank row counter
            accessory_codes_dict[code_value] = total_qty # Add value to list
            current_row += 1
   
        num_rows = len(accessory_codes_dict)

    base_scan_list = pd.DataFrame({
        "cvn_num": [collected_data["cvn_num"]] * num_rows,
        "chronological_num": list(range(1, num_rows + 1)),  # Add chronological number
        "barcode": [None] * num_rows, # Empty col for now, will be filled in later with xlookup
        "accessory_code": list(accessory_codes_dict.keys()),
        "vehicles_qty": list(accessory_codes_dict.values())
    })
    
    return base_scan_list


# Function to get the barcode by mapping item code
def xlookup_accessories(base_scan_list: pd.DataFrame) -> pd.DataFrame:
    '''
    Performs xlookp of the item code to find barcode and adds it to the dataframe
   
    Args
    scan_list (pd.DataFrame): base DataFrame containing the accessory codes to be looked up.
   
    Returns
    scan_list (pd.DataFrame): Updated DataFrame with the barcode information added.
   
    '''
    # Load accessories database file
    accessories = pd.read_csv(ACCESSORIES_DB, 
                              sep=';', 
                              encoding='latin1')

    accessories.columns = ['barcode', 'item_code', 'description', 'drop_this', 'barcode_2'] # Add colum names to be able to work with the data
    accessories = accessories.drop(columns=['drop_this', 'barcode_2']) # Drop the empty col
    accessories = accessories.dropna(subset=['item_code']) # Drop rows with empty item_code

    # Drop rows with duplicated item_codes, as this prevents the xlookup from working
    accessories = accessories.drop_duplicates(subset=['item_code'])

    # Now that the accessories database is imported and cleaned, perform xlookup
    base_scan_list['barcode'] = base_scan_list['accessory_code'].astype(str).map(accessories.set_index('item_code')['barcode']).fillna(base_scan_list['accessory_code'])
   
    base_scan_list['barcode'] = base_scan_list['barcode'].astype(str).str.replace('.0', '', regex=False)  # Remove .0 from barcodes

    return base_scan_list


def xlookup_items(scan_list: pd.DataFrame) -> pd.DataFrame:
    '''
    Performs xlookp of the barcode to find code_description and adds it to the dataframe
   
    Args
    scan_list (pd.DataFrame): base DataFrame containing the barcodes to be looked up.
   
    Returns
    scan_list (pd.DataFrame): Updated DataFrame with the item description information added.
   
    '''

    # Load accessories database file 
    items = pd.read_csv(
        ITEMS_DB, sep=';', encoding='latin1', # Remove for streamlit and use the line below instead
        # './make_scan_list/ITEMS_DATABASE.csv', sep=';', encoding='latin1', # Path for streamlit
        header = None,                      
        decimal=','
    )

    items.columns = ['barcode', 'full_description'] # Add colum names to be able to work with the data
    items = items.dropna(subset=['full_description']) # Drop rows with empty item_code
    items.drop_duplicates(subset=['barcode'], keep='first', inplace=True) # Drop duplicate barcodes

    # Now that the accessories database if imported, perform xlookup
    scan_list['barcode'] = scan_list['barcode'].astype(str)  # Ensure barcode is string for mapping
    scan_list['accessory_code'] = scan_list['barcode'].map(items.set_index('barcode')['full_description']) # Perform xlookup   

    ref_num = scan_list.iloc[0,0].split()[-4:][0]
    ref_num = f'{ref_num[:3]}-{ref_num[-4:]}'

    return scan_list, ref_num